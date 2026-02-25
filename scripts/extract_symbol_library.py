"""Extract symbol library data and images from the Excel file."""
import json
import os
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XlImage

EXCEL_PATH = "/Users/ajayjasti/Documents/Copy of Symbol Library Export.xlsm"
OUTPUT_DIR = Path("data/symbol_library")
IMAGES_DIR = OUTPUT_DIR / "images"
DB_PATH = OUTPUT_DIR / "symbol_library.json"

def extract():
    os.makedirs(IMAGES_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Symbols Library"]

    # --- Extract embedded images ---
    print(f"Found {len(ws._images)} embedded images")
    image_map = {}  # row -> list of image data
    for idx, img in enumerate(ws._images):
        try:
            # Get the anchor cell
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                row = anchor._from.row + 1  # 0-indexed to 1-indexed
                col = anchor._from.col  # 0=A, 1=B
            elif hasattr(anchor, 'row'):
                row = anchor.row
                col = anchor.col
            else:
                row = idx + 5
                col = 0

            # Save image
            img_data = img._data()
            ext = "png"
            if img_data[:3] == b'\xff\xd8\xff':
                ext = "jpg"
            elif img_data[:4] == b'\x89PNG':
                ext = "png"
            elif img_data[:4] == b'GIF8':
                ext = "gif"

            col_label = "thumb" if col == 0 else "std_thumb" if col == 1 else f"col{col}"
            fname = f"row{row}_{col_label}_{idx}.{ext}"
            img_path = IMAGES_DIR / fname
            with open(img_path, "wb") as f:
                f.write(img_data)

            if row not in image_map:
                image_map[row] = []
            image_map[row].append({
                "col": col,
                "col_label": col_label,
                "filename": fname,
            })
        except Exception as e:
            print(f"  Warning: Could not extract image {idx}: {e}")

    print(f"Extracted images for {len(image_map)} rows")

    # --- Extract symbol metadata ---
    symbols = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True), start=5):
        name = row[2]
        classification = row[3]
        status = row[4]
        pkg_text = row[5]
        ifu_text = row[6]
        sme_function = row[7]
        rev_history = row[8]
        regulations = row[9]
        notes = row[10]
        thumb_file = row[11]
        ai_file = row[12]
        std_ai_file = row[13]
        std_thumb_file = row[14]

        if not name and not pkg_text:
            continue

        # Get extracted image filenames for this row
        row_images = image_map.get(row_idx, [])
        thumb_extracted = [i["filename"] for i in row_images if i["col_label"] == "thumb"]
        std_thumb_extracted = [i["filename"] for i in row_images if i["col_label"] == "std_thumb"]

        entry = {
            "row": row_idx,
            "name": str(name) if name else "",
            "classification": str(classification) if classification else "",
            "status": str(status) if status else "",
            "pkg_text": str(pkg_text) if pkg_text else "",
            "ifu_text": str(ifu_text) if ifu_text else "",
            "sme_function": str(sme_function) if sme_function else "",
            "rev_history": str(rev_history) if rev_history else "",
            "regulations": str(regulations)[:500] if regulations else "",
            "notes": str(notes)[:500] if notes else "",
            "thumb_file_ref": str(thumb_file) if thumb_file else "",
            "ai_file_ref": str(ai_file) if ai_file else "",
            "std_thumb_file_ref": str(std_thumb_file) if std_thumb_file else "",
            "thumb_images": thumb_extracted,
            "std_thumb_images": std_thumb_extracted,
        }
        symbols.append(entry)

    # Write JSON database
    db = {
        "source": os.path.basename(EXCEL_PATH),
        "extracted_at": "2026-02-25",
        "total_symbols": len(symbols),
        "total_images": sum(len(v) for v in image_map.values()),
        "symbols": symbols,
    }

    with open(DB_PATH, "w") as f:
        json.dump(db, f, indent=2, ensure_ascii=False)

    print(f"\nExported {len(symbols)} symbols to {DB_PATH}")
    print(f"Extracted {db['total_images']} images to {IMAGES_DIR}")

    # Stats
    by_class = {}
    for s in symbols:
        c = s["classification"]
        by_class[c] = by_class.get(c, 0) + 1
    print("\nBy classification:")
    for k, v in sorted(by_class.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")

    iso_count = sum(1 for s in symbols if "ISO 15223" in s["regulations"])
    print(f"\nISO 15223 referenced: {iso_count}")

if __name__ == "__main__":
    extract()
