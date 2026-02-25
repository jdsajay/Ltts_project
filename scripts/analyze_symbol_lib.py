"""Quick script to analyze the symbol library Excel file."""
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(
    "/Users/ajayjasti/Documents/Copy of Symbol Library Export.xlsm",
    data_only=True,
)
ws = wb["Symbols Library"]

classifications = Counter()
statuses = Counter()
standard_symbols = []
all_symbols = []

for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    # Columns: A=thumb_img, B=std_thumb_img, C=name, D=classification,
    # E=status, F=pkg_text, G=ifu_text, H=sme_function, I=rev_history,
    # J=regulations, K=notes, L=thumb_file, M=ai_file, N=std_ai_file, O=std_thumb_file
    name = row[2]
    classification = row[3]
    status = row[4]
    pkg_text = row[5]
    ifu_text = row[6]
    sme = row[7]
    regs = row[9]
    thumb_file = row[11]
    std_thumb_file = row[14]

    if not name and not pkg_text:
        continue

    if classification:
        classifications[classification] += 1
    if status:
        statuses[status] += 1

    entry = {
        "name": name,
        "classification": classification,
        "status": status,
        "pkg_text": pkg_text,
        "ifu_text": ifu_text,
        "sme": sme,
        "regs": str(regs)[:200] if regs else "",
        "thumb_file": thumb_file,
        "std_thumb_file": std_thumb_file,
    }
    all_symbols.append(entry)

    if classification == "Standard":
        standard_symbols.append(entry)

print("=== Classifications ===")
for k, v in classifications.most_common():
    print(f"  {k}: {v}")

print(f"\n=== Statuses ===")
for k, v in statuses.most_common():
    print(f"  {k}: {v}")

print(f"\n=== Total symbols: {len(all_symbols)} ===")
print(f"=== Standard symbols: {len(standard_symbols)} ===")

print(f"\n=== Standard Symbols (first 30) ===")
for s in standard_symbols[:30]:
    print(f"  Name: {s['name']}")
    print(f"    Pkg text: {s['pkg_text']}")
    print(f"    Thumb: {s['thumb_file']}")
    print(f"    Regs: {s['regs'][:100]}")
    print()

# ISO 15223 symbols
iso_syms = [s for s in all_symbols if "15223" in s.get("regs", "")]
print(f"\n=== ISO 15223 referenced ({len(iso_syms)}) ===")
for s in iso_syms[:25]:
    print(f"  {s['name']} | {s['pkg_text']}")

# Breast implant specific
bi_syms = [
    s for s in all_symbols
    if any(
        kw in str(s.get("pkg_text", "")).lower()
        for kw in ["implant", "breast", "gel", "silicone", "sterile", "manufacturer", "lot", "ref", "serial", "expir", "use by", "single use", "ce mark", "do not", "caution"]
    )
]
print(f"\n=== Breast implant related ({len(bi_syms)}) ===")
for s in bi_syms[:30]:
    print(f"  [{s['classification']}] {s['name']} | {s['pkg_text']}")
